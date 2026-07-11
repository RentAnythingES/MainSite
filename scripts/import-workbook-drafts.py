import argparse
from html import unescape
import json
import os
import re
import sys
import time
import uuid
from html.parser import HTMLParser
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

import openpyxl


WORKBOOK_PATH = "Product_Database_WIP.xlsx"
BUCKET = "product-images"
PILOT_PRODUCT_NAMES = {
    "Acupressure Mat",
    "Beach Wheelchair",
    "Bike Carriers - Roof - Thule Proride 598",
    "Deuter Child Carrier",
}
CATEGORY_MAP = {
    "Baby & Children": ("baby-gear", "Baby and Children"),
    "Camping": ("travel-outdoors", "Travel and Outdoors"),
    "Costumes": ("events-celebrations", "Events & Celebrations"),
    "Fitness": ("fitness-wellness", "Fitness & Wellness"),
    "Home & Living": ("home-living", "Home and Living"),
    "Mobility": ("mobility", "Mobility and Daily Aid"),
    "Mobility & Daily Aid": ("mobility", "Mobility and Daily Aid"),
    "Mobility Aid": ("mobility", "Mobility and Daily Aid"),
    "Party & Events": ("events-celebrations", "Events & Celebrations"),
    "Photography": ("photography-content", "Photography & Content"),
    "Remote Work": ("remote-work", "Work and Tech"),
    "Travel & Hobby": ("travel-outdoors", "Travel and Outdoors"),
    "Travel & Outdoors": ("travel-outdoors", "Travel and Outdoors"),
}
DRAFT_CATEGORY = ("catalogue-review", "Catalogue Review")
USER_AGENT = "RentAnythingCatalogueReview/1.0 (+https://www.rentanything.es)"
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}


class MetadataParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.image_urls = []

    def handle_starttag(self, tag, attrs):
        attributes = {key.lower(): value for key, value in attrs if value}
        if tag == "meta":
            key = (attributes.get("property") or attributes.get("name") or "").lower()
            if key in {"og:image", "twitter:image", "twitter:image:src"} and attributes.get("content"):
                self.image_urls.append(attributes["content"])
        if tag == "img":
            for attribute in ("data-zoom-image", "data-original", "data-src", "src"):
                if attributes.get(attribute):
                    self.image_urls.append(attributes[attribute])
                    break


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def slugify(value):
    normalized = clean_text(value).lower()
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return normalized[:100]


def valid_url(value):
    parsed = urlparse(clean_text(value))
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def api_request(base_url, service_role_key, path, method="GET", payload=None, raw_body=None, content_type=None):
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
    }
    body = raw_body
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
        headers["Prefer"] = "return=representation"
    if content_type:
        headers["Content-Type"] = content_type
    request = Request(f"{base_url}{path}", data=body, method=method, headers=headers)
    try:
        with urlopen(request, timeout=30) as response:
            response_body = response.read()
            return response.status, response.headers, json.loads(response_body.decode("utf-8")) if response_body else None
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase {method} {path} failed ({error.code}): {detail}") from error


def fetch_source_image(page_url):
    request = Request(page_url, headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"})
    try:
        with urlopen(request, timeout=20) as response:
            content_type = response.headers.get_content_type()
            if content_type in ALLOWED_IMAGE_TYPES:
                return page_url, response.read(), content_type
            html = response.read().decode("utf-8", errors="replace")
    except (HTTPError, URLError, TimeoutError, ValueError):
        return None, None, None

    parser = MetadataParser()
    parser.feed(html)
    decathlon_urls = re.findall(r"https?://contents\\.mediadecathlon\\.com[^\"'\\\\\s<]+", html)
    candidates = [unescape(url).replace("\\/", "/") for url in [*decathlon_urls, *parser.image_urls]]

    for candidate in dict.fromkeys(candidates):
        image_url = urljoin(page_url, candidate)
        image_request = Request(image_url, headers={"User-Agent": USER_AGENT, "Accept": "image/avif,image/webp,image/*,*/*"})
        try:
            with urlopen(image_request, timeout=30) as response:
                content_type = response.headers.get_content_type()
                if content_type not in ALLOWED_IMAGE_TYPES:
                    continue
                image = response.read()
                if len(image) > 10 * 1024 * 1024:
                    continue
                return image_url, image, content_type
        except (HTTPError, URLError, TimeoutError, ValueError):
            continue
    return None, None, None


def upload_image(base_url, service_role_key, slug, image, content_type):
    extension = "png" if "png" in content_type else "webp" if "webp" in content_type else "jpg"
    path = f"{slug}/{int(time.time() * 1000)}-{uuid.uuid4()}.{extension}"
    status, _, _ = api_request(
        base_url,
        service_role_key,
        f"/storage/v1/object/{BUCKET}/{path}",
        method="POST",
        raw_body=image,
        content_type=content_type,
    )
    if status not in {200, 201}:
        raise RuntimeError(f"Image upload failed for {slug}")
    return f"{base_url}/storage/v1/object/public/{BUCKET}/{path}"


def load_rows():
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook["Product Database"]
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    seen_names = set()

    for spreadsheet_row, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, cells))
        name = clean_text(row.get("Product Name"))
        if not name or name in PILOT_PRODUCT_NAMES or name in seen_names:
            continue
        seen_names.add(name)
        rows.append((spreadsheet_row, row))
    return rows


def make_draft(spreadsheet_row, row, category_ids, existing_by_slug):
    name = clean_text(row.get("Product Name"))
    requested_slug = slugify(name) or f"catalogue-item-{spreadsheet_row}"
    slug = requested_slug
    if slug in existing_by_slug and existing_by_slug[slug].casefold() == name.casefold():
        return None
    if slug in existing_by_slug:
        slug = f"{requested_slug}-draft-{spreadsheet_row}"
    while slug in existing_by_slug:
        if existing_by_slug[slug].casefold() == name.casefold():
            return None
        slug = f"{requested_slug}-draft-{spreadsheet_row}-{len(existing_by_slug)}"
    existing_by_slug[slug] = name

    category_key = clean_text(row.get("Category"))
    category_slug, _ = CATEGORY_MAP.get(category_key, DRAFT_CATEGORY)
    category_id = category_ids[category_slug]
    subcategory = clean_text(row.get("Sub-Category")) or category_key or "Review required"
    source_description = clean_text(row.get("OEM Description"))
    source_specs = clean_text(row.get("Sprecifications"))
    source_features = clean_text(row.get("Features"))
    in_stock = clean_text(row.get("In Stock?")).lower() == "yes"

    specs = {
        "Import review": "Confirm product details, physical stock, pricing, and media-use approval before activation.",
    }
    if source_specs:
        specs["Source specifications"] = source_specs
    if clean_text(row.get("Asset ID")):
        specs["Asset ID"] = clean_text(row.get("Asset ID"))

    features = [feature.strip() for feature in re.split(r"[\n;•]", source_features) if feature.strip()]
    return {
        "spreadsheet_row": spreadsheet_row,
        "source_url": clean_text(row.get("Link")),
        "product": {
            "slug": slug,
            "name": name,
            "brand": clean_text(row.get("Brand")) or "Brand to confirm",
            "description": source_description or "Catalogue details are being reviewed before this product can be published.",
            "emoji": "📦",
            "image_url": None,
            "category_id": category_id,
            "subcategory": subcategory,
            "subcategory_slug": slugify(subcategory) or "review-required",
            "city": "valencia",
            "stock_total": 1 if in_stock else 0,
            "stock_available": 1 if in_stock else 0,
            "is_active": False,
            "features": features,
            "specs": specs,
        },
    }


def ensure_categories(base_url, service_role_key):
    _, _, categories = api_request(base_url, service_role_key, "/rest/v1/categories?select=id,slug,name&order=sort_order")
    category_ids = {category["slug"]: category["id"] for category in categories}
    next_sort_order = len(categories)

    for category_slug, category_name in set(CATEGORY_MAP.values()) | {DRAFT_CATEGORY}:
        if category_slug in category_ids:
            continue
        _, _, created = api_request(
            base_url,
            service_role_key,
            "/rest/v1/categories",
            method="POST",
            payload={
                "slug": category_slug,
                "name": category_name,
                "emoji": "📦",
                "description": "Draft catalogue category. No customer-facing products are active yet.",
                "sort_order": next_sort_order,
            },
        )
        category_ids[category_slug] = created[0]["id"]
        next_sort_order += 1
    return category_ids


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not service_role_key:
        raise RuntimeError("Supabase environment variables are required")

    all_rows = load_rows()
    selected_rows = all_rows[args.offset:args.offset + args.limit]
    _, _, existing = api_request(base_url, service_role_key, "/rest/v1/products?select=slug,name")
    existing_by_slug = {product["slug"]: product["name"] for product in existing}

    if args.dry_run:
        known_categories = {CATEGORY_MAP.get(clean_text(row.get("Category")), DRAFT_CATEGORY)[0] for _, row in selected_rows}
        print(json.dumps({"total_rows": len(all_rows), "selected_rows": len(selected_rows), "categories": sorted(known_categories)}, indent=2))
        return

    category_ids = ensure_categories(base_url, service_role_key)
    results = []
    for spreadsheet_row, row in selected_rows:
        draft = make_draft(spreadsheet_row, row, category_ids, existing_by_slug)
        if draft is None:
            results.append({"row": spreadsheet_row, "skipped": "already imported"})
            continue
        source_url = draft["source_url"]
        image_source, image, content_type = (None, None, None)
        if valid_url(source_url):
            image_source, image, content_type = fetch_source_image(source_url)
        if image:
            draft["product"]["image_url"] = upload_image(base_url, service_role_key, draft["product"]["slug"], image, content_type)

        _, _, created = api_request(
            base_url,
            service_role_key,
            "/rest/v1/products",
            method="POST",
            payload=draft["product"],
        )
        api_request(
            base_url,
            service_role_key,
            "/rest/v1/pricing_tiers",
            method="POST",
            payload=[{"product_id": created[0]["id"], "min_days": 1, "per_day_cents": 0}],
        )
        results.append({
            "row": spreadsheet_row,
            "slug": draft["product"]["slug"],
            "image_copied": bool(image),
            "source_url": source_url if valid_url(source_url) else None,
        })

    imported_rows = [item for item in results if "slug" in item]
    print(json.dumps({"imported": len(imported_rows), "skipped": len(results) - len(imported_rows), "with_copied_image": sum(item["image_copied"] for item in imported_rows), "rows": results}, indent=2))


if __name__ == "__main__":
    main()
